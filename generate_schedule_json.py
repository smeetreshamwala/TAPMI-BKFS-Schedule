from __future__ import annotations

import json
import re
import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parent
APP_DIR = ROOT_DIR.parent / "attendance_tracker_app"
DB_PATH = APP_DIR / "attendance.db"
ROSTER_WORKBOOK_PATH = APP_DIR / "List of students.xlsx"
SCHEDULE_OUTPUT_PATH = ROOT_DIR / "schedule.json"
SCHEDULE_EMBEDDED_OUTPUT_PATH = ROOT_DIR / "schedule-data.js"
STUDENTS_OUTPUT_PATH = ROOT_DIR / "students.json"
STUDENTS_EMBEDDED_OUTPUT_PATH = ROOT_DIR / "students-data.js"

TIMELINE_SLOTS = [
    {"label": "Slot 1", "start_time": "08:45", "end_time": "10:00", "kind": "lecture"},
    {"label": "Slot 2", "start_time": "10:15", "end_time": "11:30", "kind": "lecture"},
    {"label": "Slot 3", "start_time": "11:45", "end_time": "13:00", "kind": "lecture"},
    {"label": "Lunch Time", "start_time": "13:00", "end_time": "14:30", "kind": "lunch"},
    {"label": "Slot 4", "start_time": "14:30", "end_time": "15:45", "kind": "lecture"},
    {"label": "Slot 5", "start_time": "16:00", "end_time": "17:15", "kind": "lecture"},
    {"label": "Slot 6", "start_time": "17:45", "end_time": "19:00", "kind": "lecture"},
    {"label": "Slot 7", "start_time": "19:15", "end_time": "20:30", "kind": "lecture"},
]

STUDENT_CODE_PATTERN = re.compile(r"^25B\d{3}$", re.IGNORECASE)


def load_subjects(connection: sqlite3.Connection) -> dict[int, dict[str, Any]]:
    connection.row_factory = sqlite3.Row
    rows = connection.execute(
        """
        SELECT
            id,
            name,
            code,
            credits,
            expected_lecture_count,
            faculty,
            start_date,
            end_date,
            location,
            lecture_type,
            batch_division,
            schedule_summary
        FROM subjects
        ORDER BY name
        """
    ).fetchall()

    subjects: dict[int, dict[str, Any]] = {}
    for row in rows:
        subjects[row["id"]] = {
            "id": row["id"],
            "name": row["name"],
            "code": row["code"] or "",
            "credits": row["credits"],
            "expected_lecture_count": row["expected_lecture_count"],
            "faculty": row["faculty"] or "",
            "start_date": row["start_date"],
            "end_date": row["end_date"],
            "location": row["location"] or "",
            "lecture_type": row["lecture_type"] or "",
            "batch_division": row["batch_division"] or "",
            "schedule_summary": row["schedule_summary"] or "",
        }
    return subjects


def slot_label_for(start_time: str) -> str:
    for slot in TIMELINE_SLOTS:
        if slot["start_time"] == start_time:
            return slot["label"]
    return "Custom Slot"


def mandatory_label(is_first: int, is_last: int) -> str:
    if is_first:
        return "Mandatory - First Lecture"
    if is_last:
        return "Mandatory - Last Lecture"
    return ""


def load_lecture_entries(connection: sqlite3.Connection, subjects: dict[int, dict[str, Any]]) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT
            id,
            subject_id,
            lecture_number,
            lecture_date,
            start_time,
            end_time,
            location,
            lecture_type,
            faculty,
            batch_division,
            is_mandatory_first,
            is_mandatory_last,
            is_cancelled,
            is_extra,
            stable_id,
            source_sheet,
            source_cell,
            day_name,
            day_type,
            faculty_abbreviation,
            attendance_tracked,
            schedule_status
        FROM lectures
        ORDER BY lecture_date, start_time, subject_id, lecture_number
        """
    ).fetchall()

    entries: list[dict[str, Any]] = []
    for row in rows:
        subject = subjects[row["subject_id"]]
        entries.append(
            {
                "id": f"lecture-{row['id']}",
                "raw_id": row["id"],
                "type": "lecture",
                "date": row["lecture_date"],
                "day_name": row["day_name"] or datetime.fromisoformat(row["lecture_date"]).strftime("%A"),
                "day_type": row["day_type"] or "",
                "start_time": row["start_time"],
                "end_time": row["end_time"],
                "slot_label": slot_label_for(row["start_time"]),
                "title": subject["name"],
                "subject": subject,
                "lecture_number": row["lecture_number"],
                "faculty": row["faculty"] or subject["faculty"],
                "faculty_abbreviation": row["faculty_abbreviation"] or "",
                "location": row["location"] or subject["location"],
                "lecture_type": row["lecture_type"] or subject["lecture_type"],
                "batch_division": row["batch_division"] or subject["batch_division"],
                "attendance_tracked": bool(row["attendance_tracked"]),
                "schedule_status": row["schedule_status"] or "confirmed",
                "is_cancelled": bool(row["is_cancelled"]),
                "is_extra": bool(row["is_extra"]),
                "mandatory_label": mandatory_label(row["is_mandatory_first"], row["is_mandatory_last"]),
                "stable_id": row["stable_id"] or "",
                "source": {
                    "sheet": row["source_sheet"] or "",
                    "cell": row["source_cell"] or "",
                },
            }
        )
    return entries


def load_special_event_entries(connection: sqlite3.Connection, subjects: dict[int, dict[str, Any]]) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT
            id,
            stable_id,
            title,
            event_date,
            start_time,
            end_time,
            day_name,
            day_type,
            venue,
            source_sheet,
            source_cell,
            attendance_tracked,
            event_category,
            subject_id,
            audience_scope,
            audience_value,
            notes
        FROM special_events
        ORDER BY event_date, start_time, title
        """
    ).fetchall()

    entries: list[dict[str, Any]] = []
    for row in rows:
        linked_subject = subjects.get(row["subject_id"]) if row["subject_id"] else None
        entries.append(
            {
                "id": f"event-{row['id']}",
                "raw_id": row["id"],
                "type": "event",
                "date": row["event_date"],
                "day_name": row["day_name"] or datetime.fromisoformat(row["event_date"]).strftime("%A"),
                "day_type": row["day_type"] or "",
                "start_time": row["start_time"],
                "end_time": row["end_time"],
                "slot_label": slot_label_for(row["start_time"]),
                "title": row["title"],
                "subject": linked_subject,
                "lecture_number": None,
                "faculty": linked_subject["faculty"] if linked_subject else "",
                "faculty_abbreviation": "",
                "location": row["venue"] or (linked_subject["location"] if linked_subject else ""),
                "lecture_type": "",
                "batch_division": "",
                "attendance_tracked": bool(row["attendance_tracked"]),
                "schedule_status": "event",
                "is_cancelled": False,
                "is_extra": False,
                "mandatory_label": "",
                "stable_id": row["stable_id"] or "",
                "event_category": row["event_category"] or "general",
                "audience_scope": row["audience_scope"] or "all",
                "audience_value": row["audience_value"] or "",
                "notes": row["notes"] or "",
                "source": {
                    "sheet": row["source_sheet"] or "",
                    "cell": row["source_cell"] or "",
                },
            }
        )
    return entries


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def resolve_header_index(header_map: dict[str, int], *keys: str) -> int | None:
    for key in keys:
        if key in header_map:
            return header_map[key]
    return None


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.strip().lower()).strip("-")
    return slug or "bkfs-core"


def format_birthday(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None

    for pattern in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return text


def load_student_directory(subjects: dict[int, dict[str, Any]]) -> dict[str, Any]:
    if not ROSTER_WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Could not find roster workbook at {ROSTER_WORKBOOK_PATH}")

    workbook = load_workbook(ROSTER_WORKBOOK_PATH, data_only=True)
    worksheet = workbook["List of Students"] if "List of Students" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Roster workbook is empty.")

    header_map = {normalize_header(header): index for index, header in enumerate(rows[0]) if header not in (None, "")}
    code_index = resolve_header_index(header_map, "regno", "regno.")
    roll_index = resolve_header_index(header_map, "rollno", "rollno.")
    name_index = resolve_header_index(header_map, "fullnameasperinstituterecord", "fullname", "name")
    birthday_index = resolve_header_index(header_map, "birthday", "dateofbirth", "dob")
    section_index = resolve_header_index(header_map, "section", "track", "specialization", "batch")

    if code_index is None or name_index is None:
        raise ValueError("Roster workbook must include the student code and full name columns.")

    all_course_ids = [subject["id"] for subject in sorted(subjects.values(), key=lambda item: (item["name"], item["code"]))]
    students: list[dict[str, Any]] = []
    section_names: set[str] = set()

    for row in rows[1:]:
        if not row:
            continue
        student_code = str(row[code_index] or "").strip().upper()
        student_name = str(row[name_index] or "").strip()
        if not student_code or not student_name or not STUDENT_CODE_PATTERN.match(student_code):
            continue

        section_name = str(row[section_index] or "").strip() if section_index is not None and row[section_index] not in (None, "") else "BKFS Core"
        section_id = slugify(section_name)
        section_names.add(section_name)

        students.append(
            {
                "id": student_code.lower(),
                "student_code": student_code,
                "roll_no": str(row[roll_index]).strip() if roll_index is not None and row[roll_index] not in (None, "") else "",
                "name": student_name.title(),
                "birthday": format_birthday(row[birthday_index]) if birthday_index is not None else None,
                "section_id": section_id,
                "section_name": section_name,
                "course_ids": all_course_ids,
                "tags": [],
            }
        )

    students.sort(key=lambda item: item["student_code"])
    sections = [
        {
            "id": slugify(section_name),
            "name": section_name,
            "term": "Term 4",
            "notes": "",
        }
        for section_name in sorted(section_names)
    ] or [
        {
            "id": "bkfs-core",
            "name": "BKFS Core",
            "term": "Term 4",
            "notes": "",
        }
    ]

    default_student_code = "25B147" if any(student["student_code"] == "25B147" for student in students) else (students[0]["student_code"] if students else None)

    return {
        "generated_at": datetime.now().replace(second=0, microsecond=0).isoformat(timespec="minutes"),
        "program": "TAPMI BKFS 25-27",
        "source": {
            "type": "excel",
            "workbook_path": str(ROSTER_WORKBOOK_PATH.relative_to(ROOT_DIR.parent)).replace("\\", "/"),
            "sheet_name": worksheet.title,
            "student_count": len(students),
        },
        "default_student_code": default_student_code,
        "sections": sections,
        "students": students,
    }


def main() -> None:
    if not DB_PATH.exists():
        raise FileNotFoundError(f"Could not find attendance database at {DB_PATH}")

    connection = sqlite3.connect(DB_PATH)
    try:
        subjects = load_subjects(connection)
        lectures = load_lecture_entries(connection, subjects)
        events = load_special_event_entries(connection, subjects)
        entries = sorted(lectures + events, key=lambda item: (item["date"], item["start_time"], item["title"]))
        all_dates = [entry["date"] for entry in entries]

        schedule_payload = {
            "generated_at": datetime.now().replace(second=0, microsecond=0).isoformat(timespec="minutes"),
            "source": {
                "type": "sqlite",
                "database_path": str(DB_PATH.relative_to(ROOT_DIR.parent)).replace("\\", "/"),
                "lecture_count": len(lectures),
                "event_count": len(events),
                "subject_count": len(subjects),
            },
            "coverage": {
                "start_date": min(all_dates) if all_dates else None,
                "end_date": max(all_dates) if all_dates else None,
                "months": sorted({entry["date"][:7] for entry in entries}),
            },
            "slots": TIMELINE_SLOTS,
            "subjects": sorted(subjects.values(), key=lambda item: (item["name"], item["code"])),
            "entries": entries,
        }

        students_payload = load_student_directory(subjects)

        schedule_json_text = json.dumps(schedule_payload, indent=2, ensure_ascii=False)
        students_json_text = json.dumps(students_payload, indent=2, ensure_ascii=False)

        SCHEDULE_OUTPUT_PATH.write_text(schedule_json_text, encoding="utf-8")
        SCHEDULE_EMBEDDED_OUTPUT_PATH.write_text(
            f"window.__STATIC_SCHEDULE_DATA__ = {schedule_json_text};\n",
            encoding="utf-8",
        )
        STUDENTS_OUTPUT_PATH.write_text(students_json_text, encoding="utf-8")
        STUDENTS_EMBEDDED_OUTPUT_PATH.write_text(
            f"window.__STUDENT_DIRECTORY_DATA__ = {students_json_text};\n",
            encoding="utf-8",
        )

        print(f"Wrote {SCHEDULE_OUTPUT_PATH}")
        print(f"Wrote {SCHEDULE_EMBEDDED_OUTPUT_PATH}")
        print(f"Wrote {STUDENTS_OUTPUT_PATH}")
        print(f"Wrote {STUDENTS_EMBEDDED_OUTPUT_PATH}")
        print(
            f"Lectures: {len(lectures)} | Events: {len(events)} | Subjects: {len(subjects)} | Students: {students_payload['source']['student_count']}"
        )
    finally:
        connection.close()


if __name__ == "__main__":
    main()
